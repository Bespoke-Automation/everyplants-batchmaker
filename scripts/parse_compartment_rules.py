#!/usr/bin/env python3
"""
Parse compartment rules from Excel file and generate SQL INSERT statements.

Reads Sheet 2 ("Compartimenten") from the EveryPlants packaging module Excel,
parses all rules with their operators (EN/OF/ALTERNATIEF), and generates
SQL INSERT statements for the compartment_rules table.
"""

import re
import openpyxl

# ============================================================================
# Configuration
# ============================================================================

EXCEL_PATH = "/Users/kennylipman/Downloads/Kopie van Verpakkingsmodule basis.xlsx"
OUTPUT_PATH = "/Users/kennylipman/everyplants-batchmaker/scripts/compartment_rules.sql"
SHEET_NAME = "Compartimenten"
MAX_ROW_SCAN = 60
EMPTY_ROW_THRESHOLD = 5

COMPARTMENT_STARTS = [2, 6, 10, 14, 18, 22, 26, 30, 34, 38, 42, 46, 50, 54]

PACKAGING_UUIDS = {
    "C - Oppotten P22 - P40": "cf854272-3e12-4aa5-b911-6740eff617c0",
    "C - Oppotten P41 - P65": "f7b52ae0-85d6-45a1-a7c5-1ab1fa62bcc7",
    "C - Oppotten P66 - P80": "6d13a3c5-c2f8-47e9-8aa1-74047ba9b221",
    "C - Oppotten P81 - P100": "4efdd01e-8752-431b-ab6e-13480bd1142d",
    "C - Open Doos": "1ff5df73-3ab8-4690-ba09-2edd60f6f5a6",
    "C - Vouwdoos 160cm": "4450792b-ca52-4918-8cd6-bf0227bc6259",
    "C - Vouwdoos 130cm": "5b0068d5-7a6d-440a-81ac-8fcfef765aef",
    "C - Vouwdoos 180cm": "8cc308d5-a1ae-4b62-bdef-ff96740def2f",
    "C - Vouwdoos 100cm": "1865c42c-a3d2-474a-84ed-671a6bdc373c",
    "C - Eurodoos 60": "dd1d9a20-999e-44ce-a15a-69f2a6962c70",
    "C - Eurodoos 40": "d314a8af-4841-4f84-889a-e6ab7d428fcd",
    "C - Surprise box": "8f149e7d-0503-49ea-b37a-29deb3748004",
    "C - Kokerdoos 2x P12": "00ecc6db-0035-48a6-bfb3-d6f38e6d0c06",
    "C - 2x Kokerdoos 100cm (1 verzendlabel)": "2d6aae88-a2e8-498a-bdb8-136526c879b4",
}

SHIPPING_UNIT_UUIDS = {
    "PLANT | P10,5 - P14 | H0 - H60": "8c8fa4ba-0aa5-46e0-a1da-b122847f2b2b",
    "PLANT | P10,5 - P16 | H40 - H100": "fe0e1b2a-2fd4-4507-8293-2f1eaa812387",
    "PLANT | P17 - P21 | H0 - H100": "3548b11e-c081-46f5-aa2b-4a8148844dda",
    "PLANT | P22 - P30 | H0 - H100": "3cb78abe-0c5b-460e-96d9-d40d01408c66",
    "PLANT | P17 - P18 | H100 - H130": "79356b55-e65c-4df7-af2e-a0bc622e1c37",
    "PLANT | P19 - P24 | H100 - H130": "b32311ad-dd64-4c1f-8da1-431293b645e9",
    "PLANT | P25 - P30 | H100 - H130": "2ed5cfce-c14c-4423-a85f-a67d8485cb4b",
    "PLANT | P19 - P30 | H130 - H160": "db5569e8-b76e-492b-aaf0-0e4bf82ad130",
    "PLANT | P19 - P34 | H0 - H180": "c1c8d564-dabf-4738-8bbd-c27d1e1b9b3d",
    "PLANT | P35 - P50 | H0 - H300": "5049cbe5-2121-4869-b66c-9dc33339342f",
    "PLANT | P51 - P70 | H0 - H300": "e35ce3b7-7d39-44f3-8324-a50c2dea07fe",
    "PLANT | P71 - P80 | H0 - H300": "7f452d22-fac7-49fd-ae5d-253368e46fed",
    "PLANT | P81 - P100 | H0 - H300": "2797f20a-63e3-44fe-a8b5-b1b96641ea11",
    "POT | P17 - P18": "38cb2fad-d7bf-4084-b8f3-4cffc77afcda",
    "POT | P19 - P21": "7efa124a-77c4-4092-92f8-0491b7a5b093",
    "POT | P22 - P24": "baf62c8e-9017-4d45-9436-bf87bbb17fd9",
    "POT | P25 - P30": "2e25c709-25e0-4afe-9524-6a651a5c1114",
    "POT | P31 - P34": "4e16d4c4-33db-4522-b4d0-0d4ffd95ab4c",
    "POT | P35 - P50 (breekbaar)": "2a0ad4ec-b6ec-43e0-8f4b-55887f2d5809",
    "POT | P35 - P50 (niet breekbaar)": "8035683c-29e3-4069-8e15-dd5ce7170b46",
    "POT | P51 - P70": "9f188583-6260-4220-aa7d-00b58a579bf7",
    "POT | P71 - P100": "fb0a7503-1742-45a1-b43f-f474c2d51b09",
    "POT | Mand + Accessoires": "65176610-ace1-48ce-8c20-f7e297ddb221",
    "POT | P10,5 - P16": "f3f485ed-ed91-4e8c-98eb-1af00bbeb44a",
    "POT + PLANT | P17 - P21 | H0 - H100": "cbcf5dbd-532a-4ebf-89f4-ec8fe5437789",
    "POT + PLANT | P22 - P30 | H0 - H100": "ddf9e7fa-f6c5-4021-9e8f-592ce1d12471",
    "POT + PLANT | P17 - P18 | H100 - H130": "e65d3e97-cf19-4407-9349-cfaf6a04c89e",
    "POT + PLANT | P19 - P24 | H100 - H130": "22550be1-6334-4209-9ba0-945486b73748",
    "POT + PLANT | P24 - P30 | H100 - H130": "b45df494-9193-4d5f-975a-d139d8ec2f12",
    "POT + PLANT | P19 - P30 | H130 - H160": "c765206c-ffa2-4499-abe2-6da4f0e0b85b",
    "POT+PLANT | P19 - P34 | H0 - H180": "31288785-bf31-4417-bc29-d93eae7a9e8f",
    "POT+PLANT | P34 - P50 | H0 - H300": "abe7ecd6-9cce-4731-9181-741e7437eb94",
    "POT+PLANT | P50 - P70 | H0 - H300": "2e9a80c0-2b42-4ac4-84c5-01cec4bcdea0",
    "POT+PLANT | P70 - P100 | H0 - H300": "c646d39e-af9e-428f-a1d6-44e3fd253538",
    "POT + PLANT | P10.5 - P16 | H0 - H100": "e2201adb-3696-458e-a041-ab3837760fa8",
    "Oppotten P22 - P40": "d0d4b2aa-8a47-4f8b-9284-a5ed26abf213",
    "Oppotten P41 - P65": "5473209f-830e-40ae-a71f-4cdbf73be6f0",
    "Oppotten P66 - P80": "7c3c8efe-acde-4f86-beaa-1889a88c84d4",
    "Oppotten P81 - P100": "bb0be1cc-6c45-4608-87c4-93acc0f508d1",
    "PLANT | P5,5 - P10 | H0 - H60": "52ca0433-4879-4bff-8007-b019012fc8d7",
    "PLANT | P15 - P16 | H0 - H40": "c4283ee7-4232-4e87-813d-4fddf30d5386",
    "POT | P5,5 - P10": "4aded71a-1cda-47f8-b292-db39645633bc",
    "PLANT | P24 - P35 | H0 - H300": "5015ed85-ba97-4c06-b172-c234f1bcfe22",
}


def normalize_name(name):
    """Normalize a shipping unit name for fuzzy matching."""
    if not name:
        return ""
    s = name.strip()
    # Normalize spacing around pipes: "P80| H0" -> "P80 | H0"
    s = re.sub(r'\|(?!\s)', '| ', s)
    s = re.sub(r'(?<!\s)\|', ' |', s)
    # Normalize multiple spaces
    s = re.sub(r'\s+', ' ', s)
    return s


def build_normalized_lookup():
    """Build lookup from normalized name -> (original_name, uuid)."""
    lookup = {}
    for name, uuid in SHIPPING_UNIT_UUIDS.items():
        normalized = normalize_name(name)
        lookup[normalized] = (name, uuid)
    return lookup


def find_shipping_unit(name, lookup):
    """Find a shipping unit UUID by name with fuzzy matching."""
    if not name:
        return None, None

    normalized = normalize_name(name)

    # Direct match
    if normalized in lookup:
        return lookup[normalized]

    # Try replacing P10.5 with P10,5 and vice versa
    for old, new in [("P10.5", "P10,5"), ("P10,5", "P10.5")]:
        alt = normalized.replace(old, new)
        if alt in lookup:
            return lookup[alt]

    # Try with/without spaces around + in POT+PLANT
    for old, new in [("POT + PLANT", "POT+PLANT"), ("POT+PLANT", "POT + PLANT")]:
        alt = normalized.replace(old, new)
        if alt in lookup:
            return lookup[alt]

    # Try prefixing with "POT | " for bare names like "P22 - P24"
    if not any(normalized.startswith(p) for p in ["PLANT", "POT", "BUNDEL", "Oppotten"]):
        for prefix in ["POT | ", "PLANT | "]:
            alt = prefix + normalized
            if alt in lookup:
                return lookup[alt]

    return None, None


def parse_quantity(qty_str):
    """Parse a quantity string like '1x', '2x', '10x' to integer."""
    if qty_str is None:
        return None
    s = str(qty_str).strip().lower()
    match = re.match(r'(\d+)\s*x?', s)
    if match:
        return int(match.group(1))
    return None


def read_compartment_rules(ws, start_col, compartment_name, lookup):
    """Read all rules for a single compartment."""
    rules = []
    unmatched = []
    rule_group = 1
    sort_order = 0
    empty_count = 0
    last_non_alt_index = None

    for row in range(2, MAX_ROW_SCAN + 1):
        op_cell = ws.cell(row=row, column=start_col).value
        su_cell = ws.cell(row=row, column=start_col + 1).value
        qty_cell = ws.cell(row=row, column=start_col + 2).value

        operator = str(op_cell).strip() if op_cell else None
        su_name = str(su_cell).strip() if su_cell else None
        qty_str = str(qty_cell).strip() if qty_cell else None

        if not su_name and not operator:
            empty_count += 1
            if empty_count > EMPTY_ROW_THRESHOLD:
                break
            continue
        empty_count = 0

        quantity = parse_quantity(qty_str)
        original_name, su_uuid = find_shipping_unit(su_name, lookup)
        if su_uuid is None and su_name:
            unmatched.append((compartment_name, row, su_name))

        if operator == "OF":
            rule_group += 1
            sort_order = 0
            last_non_alt_index = None
            rule = {
                "packaging_name": compartment_name,
                "rule_group": rule_group,
                "shipping_unit_name": su_name,
                "shipping_unit_original": original_name,
                "shipping_unit_id": su_uuid,
                "quantity": quantity,
                "operator": "OF",
                "alternative_for_index": None,
                "sort_order": sort_order,
                "row": row,
            }
            last_non_alt_index = len(rules)
            rules.append(rule)
            sort_order += 1

        elif operator == "EN":
            rule = {
                "packaging_name": compartment_name,
                "rule_group": rule_group,
                "shipping_unit_name": su_name,
                "shipping_unit_original": original_name,
                "shipping_unit_id": su_uuid,
                "quantity": quantity,
                "operator": "EN",
                "alternative_for_index": None,
                "sort_order": sort_order,
                "row": row,
            }
            last_non_alt_index = len(rules)
            rules.append(rule)
            sort_order += 1

        elif operator == "ALTERNATIEF":
            rule = {
                "packaging_name": compartment_name,
                "rule_group": rule_group,
                "shipping_unit_name": su_name,
                "shipping_unit_original": original_name,
                "shipping_unit_id": su_uuid,
                "quantity": quantity,
                "operator": "ALTERNATIEF",
                "alternative_for_index": last_non_alt_index,
                "sort_order": sort_order,
                "row": row,
            }
            rules.append(rule)
            sort_order += 1

        else:
            # No operator = first rule of compartment
            rule = {
                "packaging_name": compartment_name,
                "rule_group": rule_group,
                "shipping_unit_name": su_name,
                "shipping_unit_original": original_name,
                "shipping_unit_id": su_uuid,
                "quantity": quantity,
                "operator": None,
                "alternative_for_index": None,
                "sort_order": sort_order,
                "row": row,
            }
            last_non_alt_index = len(rules)
            rules.append(rule)
            sort_order += 1

    return rules, unmatched


def generate_sql(all_rules):
    """Generate SQL INSERT statements from parsed rules."""
    lines = []
    lines.append("-- ==========================================================================")
    lines.append("-- Compartment Rules INSERT statements")
    lines.append("-- Generated from: Kopie van Verpakkingsmodule basis.xlsx")
    lines.append("-- Sheet: Compartimenten")
    lines.append("-- ==========================================================================")
    lines.append("")
    lines.append("-- First, clear existing compartment_rules")
    lines.append("DELETE FROM batchmaker.compartment_rules;")
    lines.append("")
    lines.append("-- Use a DO block to handle alternative_for_id references between rules")
    lines.append("DO $$")
    lines.append("DECLARE")

    # Assign stable variable names using compartment index + rule index
    var_map = {}
    comp_idx = 0
    for compartment_name, rules in all_rules.items():
        for i, rule in enumerate(rules):
            var_name = f"v_c{comp_idx}_r{i}"
            var_map[id(rule)] = var_name
        comp_idx += 1

    # Declare variables
    comp_idx = 0
    for compartment_name, rules in all_rules.items():
        for i, rule in enumerate(rules):
            lines.append(f"  v_c{comp_idx}_r{i} uuid;")
        comp_idx += 1

    lines.append("BEGIN")
    lines.append("")

    comp_idx = 0
    for compartment_name, rules in all_rules.items():
        packaging_uuid = PACKAGING_UUIDS.get(compartment_name)
        if not packaging_uuid:
            lines.append(f"  -- WARNING: No packaging UUID found for '{compartment_name}'")
            comp_idx += 1
            continue

        lines.append(f"  -- ====== {compartment_name} ======")

        for i, rule in enumerate(rules):
            su_uuid = rule["shipping_unit_id"]
            if not su_uuid:
                su_display = rule["shipping_unit_name"] or "UNKNOWN"
                lines.append(f"  -- WARNING: Skipping unmatched shipping unit '{su_display}' (row {rule['row']})")
                comp_idx_dummy = comp_idx  # still need to handle var
                continue

            quantity = rule["quantity"]
            if quantity is None:
                quantity = 0

            operator_sql = f"'{rule['operator']}'" if rule["operator"] else "NULL"

            # Handle alternative_for_id
            alt_for_index = rule.get("alternative_for_index")
            if alt_for_index is not None:
                alt_rule = rules[alt_for_index]
                alt_for_sql = var_map[id(alt_rule)]
            else:
                alt_for_sql = "NULL"

            var_name = var_map[id(rule)]

            lines.append(
                f"  INSERT INTO batchmaker.compartment_rules "
                f"(id, packaging_id, rule_group, shipping_unit_id, quantity, operator, alternative_for_id, sort_order, is_active) "
                f"VALUES (gen_random_uuid(), '{packaging_uuid}', {rule['rule_group']}, "
                f"'{su_uuid}', {quantity}, {operator_sql}, {alt_for_sql}, {rule['sort_order']}, true) "
                f"RETURNING id INTO {var_name};"
            )

        lines.append("")
        comp_idx += 1

    lines.append("END $$;")
    lines.append("")

    return "\n".join(lines)


def main():
    print("=" * 70)
    print("Compartment Rules Parser")
    print("=" * 70)
    print()

    print(f"Loading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    print(f"Sheet: {ws.title} ({ws.max_row} rows x {ws.max_column} cols)")
    print()

    lookup = build_normalized_lookup()

    all_rules = {}
    all_unmatched = []
    total_rules = 0

    print("Parsing compartments...")
    print("-" * 70)

    for start_col in COMPARTMENT_STARTS:
        compartment_name = ws.cell(row=1, column=start_col).value
        if not compartment_name:
            compartment_name = ws.cell(row=1, column=start_col + 1).value
        if not compartment_name:
            print(f"  WARNING: No name found at col {start_col}, skipping")
            continue

        compartment_name = str(compartment_name).strip()

        rules, unmatched = read_compartment_rules(ws, start_col, compartment_name, lookup)
        all_rules[compartment_name] = rules
        all_unmatched.extend(unmatched)
        total_rules += len(rules)

        rule_groups = set(r["rule_group"] for r in rules)
        print(f"  {compartment_name}: {len(rules)} rules, {len(rule_groups)} rule group(s)")

    print("-" * 70)
    print(f"TOTAL: {total_rules} rules across {len(all_rules)} compartments")
    print()

    if all_unmatched:
        print("WARNING: Unmatched shipping units:")
        for compartment, row, name in all_unmatched:
            print(f"  - [{compartment}] Row {row}: '{name}'")
        print()
    else:
        print("All shipping units matched successfully.")
        print()

    print(f"Generating SQL...")
    sql = generate_sql(all_rules)

    with open(OUTPUT_PATH, "w") as f:
        f.write(sql)
    print(f"SQL written to: {OUTPUT_PATH}")

    insert_count = sql.count("INSERT INTO")
    warning_count = sql.count("WARNING: Skipping")
    print(f"  {insert_count} INSERT statements")
    if warning_count:
        print(f"  {warning_count} warnings (check SQL file)")

    print()
    print("Done!")


if __name__ == "__main__":
    main()
